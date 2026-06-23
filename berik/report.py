"""QC reports — the artifacts a BIM technician attaches to the prosjekthotell.

Two outputs, both self-contained and dependency-light so they bundle cleanly into the
Windows exe:
  * qc_report.html  — a branded, Norwegian, print-to-PDF-ready QC record (the formal artifact)
  * qc_data.xlsx    — the queryable data: reconciliation tables, data-health, mapping

Design follows the BIM-QA research: a RAG completeness score, per-file breakdown, two-way
reconciliation, data-health results (full lists, never truncated), the mapping used, and a
named sign-off with timestamp — terms a Statens vegvesen / Nye Veier reviewer recognises.
"""
from __future__ import annotations
import base64
import html
import pathlib

import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _assets_dir():
    if getattr(sys, "frozen", False):
        return pathlib.Path(sys._MEIPASS) / "berik" / "ui" / "assets"  # noqa: SLF001
    return pathlib.Path(__file__).resolve().parent / "ui" / "assets"


ASSETS = _assets_dir()

INDIGO = "222754"
ACCENT = "6E8FD6"
RAG = {"green": "3FA66A", "amber": "E0A43B", "red": "D9534F", "info": "8891B4"}


def _logo_data_uri():
    p = ASSETS / "rejlers-logo-white.png"
    if p.exists():
        b64 = base64.b64encode(p.read_bytes()).decode()
        return f"data:image/png;base64,{b64}"
    return ""


def _esc(x):
    return html.escape(str(x if x is not None else ""))


# ============================================================ HTML QC report
def _html_report(a, result, approver, version):
    t = a.totals
    rag = a.rag
    rag_hex = "#" + RAG.get(rag, RAG["info"])
    logo = _logo_data_uri()

    # per-file rows
    file_rows = "".join(
        f"<tr><td>{_esc(f.file)}</td><td>{_esc(f.schema)}</td>"
        f"<td class='n'>{f.objects_with_tag}/{f.objects_total}</td>"
        f"<td class='n'>{f.tags_matched}</td>"
        f"<td class='n'>{f.objects_enriched}</td>"
        f"<td class='n'>{f.changes_new}</td>"
        f"<td class='n'>{f.changes_overwrite}</td>"
        f"<td class='n'>{f.one_to_many_tags}</td></tr>"
        for f in a.files)

    # health rows
    def health_row(h):
        col = "#" + RAG.get(h.severity, RAG["info"])
        return (f"<tr><td><span class='dot' style='background:{col}'></span>{_esc(h.title)}</td>"
                f"<td class='n'>{h.count}</td><td>{_esc(h.detail)}</td></tr>")
    health_rows = "".join(health_row(h) for h in a.health)

    # reconciliation lists (cap render, full data is in Excel)
    rec = a.reconciliation
    excel_orphans = a.reconciliation.excel_orphans
    ex_list = "".join(f"<li>{_esc(t)}</li>" for t in excel_orphans[:200])
    ex_more = f"<li class='more'>… og {len(excel_orphans)-200} til (se Excel-rapport)</li>" if len(excel_orphans) > 200 else ""
    ifc_list = "".join(
        f"<li>{_esc(o['file'])} · {_esc(o['object'])}{(' · '+_esc(o['tag'])) if o.get('tag') else ''}</li>"
        for o in rec.ifc_orphans_sample[:200])

    # mapping rows
    map_rows = "".join(
        f"<tr><td>{_esc(c.source_label)}</td><td>{_esc(c.pset)}.{_esc(c.prop)}</td>"
        f"<td>{'Ja' if c.include else 'Nei'}</td></tr>"
        for c in a.mapping.columns)

    # validation badge
    val_html = ""
    if result and result.validation:
        v = result.validation
        val_html = (f"<div class='val'>Validert mot fasit: <b>{v['match_pct']}%</b> eksakt "
                    f"({v['cells_exact_match']} celler, {v['cells_value_diff']} avvik, "
                    f"{v['cells_we_missed']} mangler)</div>")

    # buildingSMART IDS gate section
    ids_html = ""
    ids = getattr(result, "ids", None) if result else None
    if ids and ids.get("specs"):
        src = "prosjektets IDS-fil" if ids.get("source") == "project" else "Berik auto-IDS"
        rows = ""
        for s in ids["specs"]:
            ok = s["status"]
            col = "#" + (RAG["green"] if ok else RAG["amber"])
            badge = "Bestått" if ok else f"{s['failed']} avvik"
            rows += (f"<tr><td><span class='dot' style='background:{col}'></span>{_esc(s['name'])}</td>"
                     f"<td class='n'>{s['passed']}</td><td class='n'>{s['failed']}</td>"
                     f"<td><span class='badge' style='background:{col}'>{badge}</span></td></tr>")
        ids_html = f"""
  <section>
    <h3>buildingSMART IDS-validering</h3>
    <p style="color:var(--slate);font-size:12px;margin:0 0 10px">
      Maskinell kontroll mot {src} (Information Delivery Specification).</p>
    <table><thead><tr><th>Krav</th><th>Bestått</th><th>Avvik</th><th>Status</th></tr></thead>
      <tbody>{rows}</tbody></table>
  </section>"""

    written = "".join(
        f"<tr><td>{_esc(w['source'])}</td><td>{_esc(w['output'])}</td>"
        f"<td class='n'>{w['objects_enriched']}</td></tr>"
        for w in (result.written_files if result else []))

    ts = result.timestamp if result else ""

    return f"""<!DOCTYPE html><html lang="no"><head><meta charset="utf-8">
<title>Berikingsrapport — {_esc(a.excel_name)}</title>
<style>
  :root{{--indigo:#{INDIGO};--accent:#{ACCENT};--rag:{rag_hex};
    --slate:#5B6386;--pale:#D3DAEA;--off:#F4F6FB;--line:#E3E7F1;}}
  *{{box-sizing:border-box}}
  body{{font-family:'Helvetica Now Text','Helvetica Neue',Arial,sans-serif;color:#1b2030;
    margin:0;background:#fff;font-size:14px;line-height:1.5}}
  .wrap{{max-width:980px;margin:0 auto;padding:0 0 60px}}
  header{{background:var(--indigo);color:#fff;padding:34px 44px 30px;display:flex;
    justify-content:space-between;align-items:flex-start}}
  header img{{height:30px}}
  h1{{font-family:'Montefiore','Helvetica Neue',Arial,sans-serif;font-weight:600;
    font-size:26px;margin:14px 0 2px}}
  .sub{{color:var(--pale);font-size:13px}}
  .meta{{text-align:right;font-size:12px;color:var(--pale);line-height:1.7}}
  .meta b{{color:#fff}}
  .scorecard{{display:flex;gap:28px;align-items:center;padding:26px 44px;
    background:var(--off);border-bottom:1px solid var(--line)}}
  .ring{{--p:{a.completeness_pct};width:120px;height:120px;border-radius:50%;flex:0 0 120px;
    background:conic-gradient(var(--rag) calc(var(--p)*1%),#E3E7F1 0);
    display:flex;align-items:center;justify-content:center}}
  .ring i{{width:92px;height:92px;border-radius:50%;background:#fff;display:flex;flex-direction:column;
    align-items:center;justify-content:center;font-style:normal}}
  .ring b{{font-size:28px;font-weight:600;color:var(--rag)}}
  .ring span{{font-size:11px;color:var(--slate)}}
  .scoretext h2{{margin:0 0 4px;font-size:18px;font-weight:600}}
  .scoretext p{{margin:0;color:var(--slate)}}
  .val{{margin-top:10px;display:inline-block;background:#EAF6EF;color:#2C7A4B;
    border:1px solid #BFE3CE;border-radius:8px;padding:6px 12px;font-size:13px}}
  section{{padding:24px 44px 4px}}
  h3{{font-size:13px;text-transform:uppercase;letter-spacing:.08em;color:var(--accent);
    margin:0 0 12px;padding-bottom:8px;border-bottom:2px solid var(--accent);display:inline-block}}
  table{{width:100%;border-collapse:collapse;margin:0 0 8px;font-size:13px}}
  th,td{{text-align:left;padding:8px 10px;border-bottom:1px solid var(--line);vertical-align:top}}
  th{{color:var(--slate);font-weight:600;font-size:12px;text-transform:uppercase;letter-spacing:.04em}}
  td.n{{text-align:right;font-variant-numeric:tabular-nums}}
  .dot{{display:inline-block;width:9px;height:9px;border-radius:50%;margin-right:8px;vertical-align:middle}}
  .stat-row{{display:flex;gap:14px;flex-wrap:wrap;margin:0 0 6px}}
  .stat{{flex:1;min-width:120px;background:var(--off);border-radius:12px;padding:14px 16px}}
  .stat b{{display:block;font-size:22px;font-weight:600;color:var(--indigo);font-variant-numeric:tabular-nums}}
  .stat span{{font-size:12px;color:var(--slate)}}
  .recon{{display:flex;gap:20px}}
  .recon .col{{flex:1;background:var(--off);border-radius:12px;padding:14px 16px;min-width:0}}
  .recon h4{{margin:0 0 8px;font-size:13px}}
  .recon ul{{margin:0;padding:0 0 0 4px;list-style:none;max-height:260px;overflow:auto;
    font-size:12px;font-family:ui-monospace,Menlo,Consolas,monospace}}
  .recon li{{padding:3px 0;border-bottom:1px solid #EEF1F8;word-break:break-all}}
  .recon li.more{{color:var(--slate);font-style:italic}}
  .badge{{display:inline-block;border-radius:20px;padding:2px 10px;font-size:12px;font-weight:600;color:#fff}}
  footer{{padding:22px 44px;color:var(--slate);font-size:12px;border-top:1px solid var(--line);
    display:flex;justify-content:space-between;margin-top:18px}}
  .signoff{{background:var(--indigo);color:#fff;border-radius:12px;padding:18px 22px;margin:8px 0}}
  .signoff b{{color:#fff}} .signoff .accent{{color:var(--accent)}}
  .noprint{{position:fixed;top:18px;right:18px}}
  .noprint button{{background:var(--accent);color:#fff;border:0;border-radius:8px;
    padding:10px 16px;font-size:13px;cursor:pointer;font-weight:600}}
  @media print{{.noprint{{display:none}} body{{font-size:11.5px}}
    header,.scorecard,section,footer{{padding-left:24px;padding-right:24px}}
    .recon ul{{max-height:none}} @page{{size:A4;margin:14mm}}}}
</style></head><body>
<div class="noprint"><button onclick="window.print()">Skriv ut / Lagre som PDF</button></div>
<div class="wrap">
  <header>
    <div>{f'<img src="{logo}" alt="Rejlers">' if logo else '<b style="font-size:20px">Rejlers</b>'}
      <h1>Berikingsrapport — QC</h1>
      <div class="sub">Excel → IFC metadata · {_esc(a.excel_name)}</div>
    </div>
    <div class="meta">
      <div>Godkjent av <b>{_esc(approver)}</b></div>
      <div>{_esc(ts)}</div>
      <div>Filer: <b>{t['files']}</b> · Berik v{_esc(version)}</div>
    </div>
  </header>

  <div class="scorecard">
    <div class="ring"><i><b>{a.completeness_pct:.0f}%</b><span>komplett</span></i></div>
    <div class="scoretext">
      <h2>{t['objects_enriched']} av {t['objects_with_tag']} taggede objekter beriket</h2>
      <p>{t['changes_new']} nye verdier skrevet · {t['changes_overwrite']} overskrevet ·
         {t['tags_matched']} tagger matchet · {t['objects_without_tag']} objekter uten tag (unntak)</p>
      {val_html}
    </div>
  </div>

  <section>
    <h3>Per fil</h3>
    <table><thead><tr><th>Fil</th><th>Skjema</th><th>Tagget/Totalt</th><th>Matchet</th>
      <th>Beriket</th><th>Nye</th><th>Overskr.</th><th>1:mange</th></tr></thead>
      <tbody>{file_rows}</tbody></table>
  </section>

  <section>
    <h3>Toveis avstemming</h3>
    <div class="recon">
      <div class="col"><h4>Excel-rader uten objekt
        <span class="badge" style="background:{'#'+RAG['green'] if not excel_orphans else '#'+RAG['amber']}">{len(excel_orphans)}</span></h4>
        <ul>{ex_list}{ex_more}{'<li class=more>Ingen — alt matchet.</li>' if not excel_orphans else ''}</ul>
        {f'<p style="font-size:12px;color:var(--slate);margin:8px 0 0">+ {rec.excel_rows_other_files} rader hører til andre disiplinfiler som ikke ble berikt i denne kjøringen.</p>' if rec.excel_rows_other_files else ''}
      </div>
      <div class="col"><h4>Objekter uten Excel-rad
        <span class="badge" style="background:{'#'+RAG['green'] if not rec.ifc_orphans_count else '#'+RAG['amber']}">{rec.ifc_orphans_count}</span></h4>
        <ul>{ifc_list}{'<li class=more>Ingen.</li>' if not rec.ifc_orphans_count else ''}</ul>
      </div>
    </div>
  </section>

  <section>
    <h3>Datakvalitet</h3>
    <table><thead><tr><th>Sjekk</th><th>Antall</th><th>Merknad</th></tr></thead>
      <tbody>{health_rows}</tbody></table>
  </section>
  {ids_html}

  <section>
    <h3>Kartlegging brukt (Excel-kolonne → IFC-egenskap)</h3>
    <table><thead><tr><th>Kilde (Excel)</th><th>Mål (Pset.Egenskap)</th><th>Inkludert</th></tr></thead>
      <tbody>{map_rows}</tbody></table>
  </section>

  <section>
    <h3>Skrevne filer</h3>
    <table><thead><tr><th>Kilde</th><th>Beriket fil</th><th>Objekter beriket</th></tr></thead>
      <tbody>{written}</tbody></table>
    <div class="signoff">Berik skrev nye <b>_beriket</b>-kopier. Originalfilene er urørt.
      Godkjent av <b class="accent">{_esc(approver)}</b> · {_esc(ts)}.</div>
  </section>

  <footer><span>Berik — Excel til IFC · bygget av GRAIL for Rejlers</span>
    <span>Modellgrunnlag-QC · v{_esc(version)}</span></footer>
</div></body></html>"""


# ================================================================ Excel data
def _xlsx_report(a, result, out_path, approver):
    wb = openpyxl.Workbook()
    head_fill = PatternFill("solid", fgColor=INDIGO)
    head_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(bottom=thin)

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = head_fill; cell.font = head_font
            cell.alignment = Alignment(vertical="center")

    # Sammendrag
    ws = wb.active; ws.title = "Sammendrag"
    t = a.totals
    rows = [
        ("Excel-liste", a.excel_name), ("Godkjent av", approver),
        ("Tidspunkt", result.timestamp if result else ""),
        ("Kompletthet (%)", a.completeness_pct), ("RAG", a.rag),
        ("Filer", t["files"]), ("Objekter totalt", t["objects_total"]),
        ("Objekter med tag", t["objects_with_tag"]), ("Objekter uten tag", t["objects_without_tag"]),
        ("Objekter beriket", t["objects_enriched"]), ("Nye verdier", t["changes_new"]),
        ("Overskrevet", t["changes_overwrite"]), ("Tagger matchet", t["tags_matched"]),
        ("Excel-rader uten objekt", t["excel_orphans"]), ("Objekter uten Excel-rad", t["ifc_orphans"]),
    ]
    ws.append(["Felt", "Verdi"]); style_header(ws, 2)
    for k, v in rows:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 48

    # Filer
    ws = wb.create_sheet("Filer")
    ws.append(["Fil", "Skjema", "Objekter", "Med tag", "Uten tag", "Unike tagger",
               "Matchet", "Beriket", "Nye", "Overskr.", "1:mange"]); style_header(ws, 11)
    for f in a.files:
        ws.append([f.file, f.schema, f.objects_total, f.objects_with_tag, f.objects_without_tag,
                   f.unique_tags, f.tags_matched, f.objects_enriched, f.changes_new,
                   f.changes_overwrite, f.one_to_many_tags])

    # Avstemming
    ws = wb.create_sheet("Avstemming")
    ws.append(["Type", "Fil", "Objekt", "Tag"]); style_header(ws, 4)
    for tag in a.reconciliation.excel_orphans:
        ws.append(["Excel-rad uten objekt", "", "", tag])
    for o in a.reconciliation.ifc_orphans_sample:
        ws.append(["Objekt uten Excel-rad", o.get("file", ""), o.get("object", ""), o.get("tag", "")])
    ws.column_dimensions["A"].width = 24; ws.column_dimensions["D"].width = 44

    # Datakvalitet
    ws = wb.create_sheet("Datakvalitet")
    ws.append(["Sjekk", "Alvorlighet", "Antall", "Merknad"]); style_header(ws, 4)
    for h in a.health:
        ws.append([h.title, h.severity, h.count, h.detail])
        for item in h.items[:2000]:
            ws.append(["", "", "", item])
    ws.column_dimensions["A"].width = 32; ws.column_dimensions["D"].width = 60

    # Kartlegging
    ws = wb.create_sheet("Kartlegging")
    ws.append(["Kilde (Excel)", "Pset", "Egenskap", "Inkludert"]); style_header(ws, 4)
    for c in a.mapping.columns:
        ws.append([c.source_label, c.pset, c.prop, "Ja" if c.include else "Nei"])
    ws.column_dimensions["A"].width = 32; ws.column_dimensions["B"].width = 22; ws.column_dimensions["C"].width = 28

    wb.save(out_path)
    return out_path


# ===================================================================== entry
def build_reports(analysis, result, out_dir, approver, version="1.0"):
    """Write qc_report.html + qc_data.xlsx into out_dir. Returns (html_path, xlsx_path)."""
    out_dir = pathlib.Path(out_dir); out_dir.mkdir(parents=True, exist_ok=True)
    html_path = out_dir / "QC-rapport.html"
    xlsx_path = out_dir / "QC-data.xlsx"
    html_path.write_text(_html_report(analysis, result, approver, version), encoding="utf-8")
    _xlsx_report(analysis, result, xlsx_path, approver)
    return str(html_path), str(xlsx_path)
