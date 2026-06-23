"""Read the Excel equipment list directly (no CSV step) and build the mapping + data.

The workbook's `CSV` tab is the upstream tool's import-format spec: row 2 carries the Pset prefix
(`0_Element::`, `1_ELE::`, `2_RDS::`), row 3 the property name, data from row 4 down,
keyed by the TFM Tag in column 0. We parse that structure into a Mapping and a
{tag: {(pset, prop): value}} dict, applying the three rules learned from ground truth:

  * blank cell -> omit the property; zero -> skip  (the upstream "blank stays blank" rule)
  * doc-link columns (D1..D5) route to the `D_Dokumentknyttinger` Pset
  * a duplicated full tag keeps the row carrying the most data (real row beats stub)
"""
from __future__ import annotations
import pathlib
import openpyxl

from .model import Mapping, ColumnSpec

KEY_PSET = "1_ELE"
KEY_PROP = "TFM Tag"

# The upstream tool routes these columns (labelled `1_ELE::` in the sheet) to a dedicated Pset.
DOC_PROPS = {"D1 Installasjon", "D2 Skjema", "D3 Topologi", "D4", "D5"}
DOC_PSET = "D_Dokumentknyttinger"

# Tabs that look like an upstream / IfcCSV import sheet, in preference order.
CANDIDATE_TABS = ("CSV", "IFC", "Export", "Eksport")

# Column that names which IFC file a row belongs to (lets us scope orphans per file).
DISCIPLINE_PROP = "Disiplinmodell"

# IFC filename suffixes that mark a pre-/post-enrichment variant of the same model.
STEM_SUFFIXES = ("_blank", "_uten_data", "_clean", "_beriket", "_beriket_grail")


def normalize_stem(name: str) -> str:
    """Normalise an IFC filename or Disiplinmodell value to a comparable model key."""
    s = str(name).strip().lower()
    if s.endswith(".ifc"):
        s = s[:-4]
    for suf in STEM_SUFFIXES:
        if s.endswith(suf):
            s = s[: -len(suf)]
            break
    return s


def clean(value):
    """Writable string, or None if the cell must be skipped (blank / zero)."""
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        return None if v in ("", "0") else v
    if isinstance(value, (int, float)):
        if value == 0:
            return None
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    s = str(value).strip()
    return s or None


def detect_tab(path: pathlib.Path) -> str:
    """Pick the mapping tab: a known name, else the first sheet whose row 2 carries `::`."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    for cand in CANDIDATE_TABS:
        if cand in names:
            return cand
    for name in names:
        ws = wb[name]
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
            if any(c and "::" in str(c) for c in row):
                return name
    return names[0]


def read_workbook(path: pathlib.Path, tab: str | None = None):
    """Return (Mapping, rows_by_tag). rows_by_tag: {tag: {(pset, prop): clean_value}}."""
    path = pathlib.Path(path)
    tab = tab or detect_tab(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if tab not in wb.sheetnames:
        raise ValueError(f"Tab '{tab}' not found. Tabs: {wb.sheetnames}")
    ws = wb[tab]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        raise ValueError(f"Tab '{tab}' has too few rows to be a mapping sheet.")
    pset_row, name_row = rows[1], rows[2]   # row 2 = Pset prefix, row 3 = property name

    columns: list[ColumnSpec] = []
    col_targets: dict[int, tuple[str, str]] = {}
    for i, raw in enumerate(pset_row):
        if raw and "::" in str(raw):
            pset = str(raw).replace("::", "").strip()
            prop = str(name_row[i]).strip() if name_row[i] is not None else ""
            if not prop:
                continue
            if prop in DOC_PROPS:
                pset = DOC_PSET
            columns.append(ColumnSpec(col_index=i, pset=pset, prop=prop,
                                      source_label=str(raw).replace("::", "").strip()))
            col_targets[i] = (pset, prop)

    mapping = Mapping(tab=tab, key_pset=KEY_PSET, key_prop=KEY_PROP, columns=columns)

    # which column carries the Disiplinmodell value (the per-row file name)
    disc_col = next((i for i, (_, prop) in col_targets.items() if prop == DISCIPLINE_PROP), None)

    rows_by_tag: dict[str, dict[tuple[str, str], str]] = {}
    tag_discipline: dict[str, str | None] = {}
    for r in rows[3:]:
        if not r or r[0] is None:
            continue
        tag = str(r[0]).strip()
        if not tag:
            continue
        props: dict[tuple[str, str], str] = {}
        for i, (pset, prop) in col_targets.items():
            if i < len(r):
                cv = clean(r[i])
                if cv is not None:
                    props[(pset, prop)] = cv
        # Duplicate full tag: the row carrying the most data wins (real row beats stub).
        if tag not in rows_by_tag or len(props) > len(rows_by_tag[tag]):
            rows_by_tag[tag] = props
            disc = clean(r[disc_col]) if (disc_col is not None and disc_col < len(r)) else None
            tag_discipline[tag] = normalize_stem(disc) if disc else None
    return mapping, rows_by_tag, tag_discipline


def raw_tag_rows(path: pathlib.Path, tab: str):
    """All TFM tags as they appear (incl. duplicates) — for duplicate-detection health check."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[tab]
    out = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r and r[0] is not None and str(r[0]).strip():
            out.append(str(r[0]).strip())
    return out
